"""
Excel report generator for flipping analysis results.

Generates a formatted .xlsx file with:
  - Sheet 1 "Ranking"      : full ranked table of all analyzed properties
  - Sheet 2 "Detalle"      : one row per property with pros, cons and renovation suggestions
  - Sheet 3 "Mercado"      : market reference summary
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models import FlippingAnalysis, MarketReference


# ── Colour palette ────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1F3864"   # dark navy
CLR_HEADER_FG   = "FFFFFF"
CLR_BUY_BG      = "C6EFCE"   # light green
CLR_BUY_FG      = "276221"
CLR_WATCH_BG    = "FFEB9C"   # light yellow
CLR_WATCH_FG    = "9C6500"
CLR_SKIP_BG     = "FFC7CE"   # light red
CLR_SKIP_FG     = "9C0006"
CLR_ALT_ROW     = "F2F7FF"   # very light blue for alternating rows
CLR_SECTION_BG  = "D9E1F2"   # section header blue
CLR_SCORE_HIGH  = "375623"   # dark green text for high scores
CLR_SCORE_MED   = "7D6608"
CLR_SCORE_LOW   = "7D1A1A"
CLR_TITLE_BG    = "2E75B6"   # blue title bar


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic)


def _border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _rec_colors(rec: str) -> tuple[str, str]:
    if rec == "COMPRAR":
        return CLR_BUY_BG, CLR_BUY_FG
    if rec == "ANALIZAR MÁS":
        return CLR_WATCH_BG, CLR_WATCH_FG
    return CLR_SKIP_BG, CLR_SKIP_FG


def _score_color(score: float) -> str:
    if score >= 7:
        return CLR_SCORE_HIGH
    if score >= 5:
        return CLR_SCORE_MED
    return CLR_SCORE_LOW


def _write_title_row(ws: Worksheet, title: str, ncols: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = _fill(CLR_TITLE_BG)
    cell.font = _font(bold=True, color="FFFFFF", size=13)
    cell.alignment = _center()
    ws.row_dimensions[1].height = 28


def _write_header(ws: Worksheet, row: int, columns: list[str]) -> None:
    for col, name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill = _fill(CLR_HEADER_BG)
        cell.font = _font(bold=True, color=CLR_HEADER_FG, size=10)
        cell.alignment = _center(wrap=True)
        cell.border = _border()
    ws.row_dimensions[row].height = 36


# ── Sheet 1: Ranking ──────────────────────────────────────────────────────────

RANKING_COLS = [
    "#", "Recomendación", "Score /10", "Barrio", "Tipo",
    "Título", "Precio (USD)", "Sup. cubierta (m²)", "USD/m²",
    "Desc. vs. mercado", "ARV estimado (USD)", "Costo reno. (USD)",
    "Ganancia est. (USD)", "ROI estimado", "Fuente", "URL",
]

RANKING_WIDTHS = [4, 14, 9, 14, 12, 40, 14, 14, 10, 14, 16, 14, 16, 10, 13, 50]


def _build_ranking_sheet(ws: Worksheet, analyses: list[FlippingAnalysis]) -> None:
    ws.title = "Ranking"
    ws.freeze_panes = "A3"

    _write_title_row(ws, "🏠 Ranking de Oportunidades de Flipping — Buenos Aires", len(RANKING_COLS))
    _write_header(ws, 2, RANKING_COLS)

    for i, a in enumerate(analyses, 1):
        p = a.property
        row = i + 2
        surface = p.surface_covered or p.surface_total

        values = [
            i,
            a.flipping_recommendation or "—",
            a.flipping_score,
            p.neighborhood,
            p.property_type.value.title(),
            p.title,
            p.price_usd,
            surface,
            p.price_per_m2,
            (a.discount_vs_market_pct / 100) if a.discount_vs_market_pct is not None else None,
            a.estimated_arv_usd,
            a.estimated_renovation_cost_usd,
            a.estimated_profit_usd,
            (a.roi_pct / 100) if a.roi_pct is not None else None,
            p.source.value,
            p.url,
        ]

        alt = i % 2 == 0
        bg_fill = _fill(CLR_ALT_ROW) if alt else _fill("FFFFFF")
        rec_bg, rec_fg = _rec_colors(a.flipping_recommendation or "")

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _border()
            cell.alignment = _left(wrap=False)

            # Recommendation column → coloured
            if col == 2:
                cell.fill = _fill(rec_bg)
                cell.font = _font(bold=True, color=rec_fg, size=10)
                cell.alignment = _center()
            # Score column
            elif col == 3:
                cell.fill = bg_fill
                cell.font = _font(bold=True, color=_score_color(a.flipping_score), size=10)
                cell.alignment = _center()
                cell.number_format = "0.0"
            # Percentage columns
            elif col in (10, 14):
                cell.fill = bg_fill
                cell.number_format = "0.0%"
                if val is not None:
                    if col == 10:  # discount
                        color = CLR_BUY_FG if val <= -0.05 else (CLR_SKIP_FG if val >= 0.1 else "000000")
                    else:          # ROI
                        color = CLR_BUY_FG if val >= 0.1 else (CLR_SKIP_FG if val < 0 else "000000")
                    cell.font = _font(bold=True, color=color, size=10)
            # Currency columns
            elif col in (7, 9, 11, 12, 13):
                cell.fill = bg_fill
                cell.number_format = '"USD "#,##0'
                cell.font = _font(size=10)
            # URL → hyperlink style
            elif col == 16:
                cell.fill = bg_fill
                cell.font = Font(color="0563C1", underline="single", size=9)
                cell.hyperlink = val or ""
            else:
                cell.fill = bg_fill
                cell.font = _font(size=10)

        ws.row_dimensions[row].height = 18

    # Column widths
    for col, width in enumerate(RANKING_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Auto-filter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(RANKING_COLS))}2"


# ── Sheet 2: Detalle ──────────────────────────────────────────────────────────

DETAIL_COLS = [
    "#", "Título", "Barrio", "Dirección", "Tipo", "Ambientes", "Dorm.",
    "Baños", "Piso", "Antigüedad", "Amenities", "Cochera", "Expensas (ARS)",
    "Fotos", "Descripción",
    "Pros", "Contras", "Sugerencias de renovación",
    "Resumen IA", "Score", "Recomendación",
]

DETAIL_WIDTHS = [4, 40, 14, 30, 12, 10, 8, 8, 8, 10, 30, 8, 14, 6, 60, 50, 50, 50, 60, 8, 14]


def _build_detail_sheet(ws: Worksheet, analyses: list[FlippingAnalysis]) -> None:
    ws.title = "Detalle Propiedades"
    ws.freeze_panes = "A3"

    _write_title_row(ws, "📋 Detalle de Propiedades Analizadas", len(DETAIL_COLS))
    _write_header(ws, 2, DETAIL_COLS)

    for i, a in enumerate(analyses, 1):
        p = a.property
        row = i + 2
        alt = i % 2 == 0
        bg_fill = _fill(CLR_ALT_ROW) if alt else _fill("FFFFFF")
        rec_bg, rec_fg = _rec_colors(a.flipping_recommendation or "")

        pros_text = "\n".join(f"✓ {x}" for x in a.pros) if a.pros else "—"
        cons_text = "\n".join(f"✗ {x}" for x in a.cons) if a.cons else "—"
        reno_text = "\n".join(f"→ {x}" for x in a.renovation_suggestions) if a.renovation_suggestions else "—"

        values = [
            i,
            p.title,
            p.neighborhood,
            p.address,
            p.property_type.value.title(),
            p.rooms,
            p.bedrooms,
            p.bathrooms,
            f"Piso {p.floor}/{p.total_floors}" if p.floor else "—",
            f"{p.antiquity_years} años" if p.antiquity_years else "—",
            ", ".join(p.amenities) if p.amenities else "—",
            "Sí" if p.parking else "No",
            p.expenses,
            p.photos_count,
            p.description[:500] if p.description else "—",
            pros_text,
            cons_text,
            reno_text,
            a.ai_summary,
            a.flipping_score,
            a.flipping_recommendation or "—",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _border()
            cell.font = _font(size=10)
            cell.alignment = _left(wrap=True)

            if col in (16, 17, 18, 19, 15):  # text blocks
                cell.fill = bg_fill
            elif col == 20:  # score
                cell.fill = bg_fill
                cell.font = _font(bold=True, color=_score_color(a.flipping_score), size=10)
                cell.alignment = _center()
            elif col == 21:  # recommendation
                cell.fill = _fill(rec_bg)
                cell.font = _font(bold=True, color=rec_fg, size=10)
                cell.alignment = _center()
            elif col == 13 and val:  # expenses
                cell.number_format = "#,##0"
                cell.fill = bg_fill
            else:
                cell.fill = bg_fill

        # Row height for wrapped text
        n_lines = max(
            len(pros_text.split("\n")),
            len(cons_text.split("\n")),
            len(reno_text.split("\n")),
        )
        ws.row_dimensions[row].height = max(40, n_lines * 15)

    for col, width in enumerate(DETAIL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Sheet 3: Mercado ──────────────────────────────────────────────────────────

def _build_market_sheet(
    ws: Worksheet,
    market_ref: MarketReference,
    total_props: int,
    neighborhood: str,
) -> None:
    ws.title = "Referencia de Mercado"

    ws.merge_cells("A1:B1")
    cell = ws.cell(row=1, column=1, value="📊 Referencia de Mercado")
    cell.fill = _fill(CLR_TITLE_BG)
    cell.font = _font(bold=True, color="FFFFFF", size=13)
    cell.alignment = _center()
    ws.row_dimensions[1].height = 28

    rows = [
        ("Barrio / zona", neighborhood),
        ("Total propiedades analizadas", total_props),
        ("Propiedades con precio/m²", market_ref.sample_count),
        ("Precio promedio (USD/m²)", market_ref.avg_price_per_m2_usd),
        ("Precio mediana (USD/m²)", market_ref.median_price_per_m2_usd),
        ("Precio mínimo (USD/m²)", market_ref.min_price_per_m2_usd),
        ("Precio máximo (USD/m²)", market_ref.max_price_per_m2_usd),
        ("Tipo de propiedad referencia", market_ref.property_type.value.title()),
    ]

    for i, (label, value) in enumerate(rows, 2):
        lbl_cell = ws.cell(row=i, column=1, value=label)
        lbl_cell.fill = _fill(CLR_SECTION_BG)
        lbl_cell.font = _font(bold=True, size=11)
        lbl_cell.border = _border()
        lbl_cell.alignment = _left()

        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.fill = _fill("FFFFFF") if i % 2 == 0 else _fill(CLR_ALT_ROW)
        val_cell.font = _font(size=11)
        val_cell.border = _border()
        val_cell.alignment = _left()
        if isinstance(value, float) and i >= 4:
            val_cell.number_format = '"USD "#,##0'

        ws.row_dimensions[i].height = 22

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


# ── Public API ────────────────────────────────────────────────────────────────

def export_excel(
    analyses: list[FlippingAnalysis],
    market_ref: MarketReference,
    neighborhood: str,
    total_props: int,
    output_path: Optional[str] = None,
) -> str:
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        nb = neighborhood.replace(" ", "_").lower()
        output_path = f"flipping_{nb}_{ts}.xlsx"

    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    ws_ranking = wb.create_sheet("Ranking")
    ws_detail  = wb.create_sheet("Detalle Propiedades")
    ws_market  = wb.create_sheet("Referencia de Mercado")

    _build_ranking_sheet(ws_ranking, analyses)
    _build_detail_sheet(ws_detail, analyses)
    _build_market_sheet(ws_market, market_ref, total_props, neighborhood)

    # Set Ranking as the active sheet on open
    wb.active = ws_ranking

    wb.save(output_path)
    return output_path
